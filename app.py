import os
import io
import json
import base64
import string
import random
import tempfile
import secrets
import urllib.request
import psycopg2
import psycopg2.extras
from functools import wraps
from flask import Flask, request, jsonify, send_file
from werkzeug.security import generate_password_hash, check_password_hash

app = Flask(__name__, static_folder='static', static_url_path='')

DATABASE_URL = os.environ.get('DATABASE_URL', 'sqlite://:memory:')

def get_db():
    conn = psycopg2.connect(DATABASE_URL)
    conn.autocommit = True
    return conn

def init_db():
    try:
        with get_db() as db:
            cur = db.cursor()
            cur.execute('''CREATE TABLE IF NOT EXISTS proyectos (
                id TEXT PRIMARY KEY,
                nombre TEXT NOT NULL,
                creado_en TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )''')
            cur.execute('''CREATE TABLE IF NOT EXISTS rendiciones (
                id SERIAL PRIMARY KEY,
                proyecto_id TEXT NOT NULL,
                nombre_persona_gasto TEXT NOT NULL,
                boleta_o_factura TEXT NOT NULL,
                empresa_emite TEXT NOT NULL,
                nro_boleta_factura TEXT,
                monto_neto REAL DEFAULT 0,
                monto_total REAL DEFAULT 0,
                fecha TEXT,
                hora TEXT,
                imagen_url TEXT,
                creado_en TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (proyecto_id) REFERENCES proyectos(id)
            )''')
            cur.execute('''ALTER TABLE rendiciones ADD COLUMN IF NOT EXISTS imagen_url TEXT''')

            cur.execute('''CREATE TABLE IF NOT EXISTS usuarios (
                id SERIAL PRIMARY KEY,
                email TEXT UNIQUE NOT NULL,
                password_hash TEXT NOT NULL,
                rol TEXT NOT NULL CHECK (rol IN ('trabajador', 'contador')),
                codigo_trabajador TEXT UNIQUE,
                creado_en TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )''')
            cur.execute('''CREATE TABLE IF NOT EXISTS sesiones (
                token TEXT PRIMARY KEY,
                usuario_id INTEGER NOT NULL,
                creado_en TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (usuario_id) REFERENCES usuarios(id)
            )''')
            cur.execute('''CREATE TABLE IF NOT EXISTS vinculaciones (
                id SERIAL PRIMARY KEY,
                contador_id INTEGER NOT NULL,
                trabajador_id INTEGER NOT NULL,
                creado_en TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (contador_id) REFERENCES usuarios(id),
                FOREIGN KEY (trabajador_id) REFERENCES usuarios(id),
                UNIQUE(contador_id, trabajador_id)
            )''')
            cur.execute('''CREATE TABLE IF NOT EXISTS rendiciones_ejecutivas (
                id SERIAL PRIMARY KEY,
                usuario_id INTEGER NOT NULL,
                tipo TEXT NOT NULL CHECK (tipo IN ('compania', 'restitucion')),
                nombre TEXT NOT NULL,
                fecha TEXT NOT NULL,
                monto_total REAL DEFAULT 0,
                estado TEXT NOT NULL DEFAULT 'activa' CHECK (estado IN ('activa', 'cerrada')),
                creado_en TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (usuario_id) REFERENCES usuarios(id)
            )''')
            cur.execute('''CREATE TABLE IF NOT EXISTS detalles_rendicion (
                id SERIAL PRIMARY KEY,
                rendicion_id INTEGER NOT NULL,
                tipo_gasto_entrada TEXT NOT NULL CHECK (tipo_gasto_entrada IN ('boleta', 'factura', 'devolucion', 'manual')),
                fecha TEXT,
                rut_emisor TEXT,
                nro_documento TEXT,
                monto_total REAL DEFAULT 0,
                monto_neto REAL DEFAULT 0,
                monto_iva REAL DEFAULT 0,
                empresa_emite TEXT,
                tipo_gasto TEXT,
                descripcion TEXT,
                imagen_url TEXT,
                creado_en TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (rendicion_id) REFERENCES rendiciones_ejecutivas(id)
            )''')
            cur.execute('''ALTER TABLE detalles_rendicion ADD COLUMN IF NOT EXISTS empresa_emite TEXT''')
            cur.execute('''ALTER TABLE detalles_rendicion ADD COLUMN IF NOT EXISTS monto_neto REAL DEFAULT 0''')
            cur.execute('''ALTER TABLE detalles_rendicion ADD COLUMN IF NOT EXISTS monto_iva REAL DEFAULT 0''')
            cur.execute('''ALTER TABLE detalles_rendicion ADD COLUMN IF NOT EXISTS centro_costo_id INTEGER''')

            cur.execute('''CREATE TABLE IF NOT EXISTS centros_costo (
                id SERIAL PRIMARY KEY,
                codigo TEXT UNIQUE NOT NULL,
                nombre TEXT NOT NULL,
                creado_en TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )''')
            cur.execute('''ALTER TABLE centros_costo ADD COLUMN IF NOT EXISTS codigo TEXT''')
            try:
                cur.execute('''ALTER TABLE centros_costo ADD CONSTRAINT centros_costo_codigo_key UNIQUE (codigo)''')
            except Exception:
                pass

            centros_default = [
                ('110101', 'Finanzas'),
                ('110102', 'Contabilidad'),
                ('110103', 'Tesorería'),
                ('110104', 'Auditoría Interna'),
                ('110105', 'Control de Gestión'),
                ('110106', 'IT (Tecnología de Información)'),
                ('110107', 'Soporte Técnico'),
                ('110108', 'Desarrollo de Software'),
                ('110109', 'Operaciones'),
                ('110110', 'Logística y Distribución'),
                ('110111', 'MTK (Marketing)'),
                ('110112', 'Ventas Nacionales'),
                ('110113', 'Ventas Internacionales'),
                ('110114', 'Servicio al Cliente'),
                ('110115', 'E-commerce'),
                ('110116', 'Recursos Humanos'),
                ('110117', 'Capacitación y Desarrollo'),
                ('110118', 'Legal y Cumplimiento'),
                ('110119', 'Compras y Abastecimiento'),
                ('110120', 'Servicios Generales'),
            ]
            for cod, nom in centros_default:
                cur.execute(
                    'INSERT INTO centros_costo (codigo, nombre) VALUES (%s, %s) ON CONFLICT (codigo) DO NOTHING',
                    (cod, nom)
                )
    except Exception as e:
        print("Error inicializando DB:", e)

def generar_id(length=8):
    chars = string.ascii_lowercase + string.digits
    return ''.join(random.choice(chars) for _ in range(length))

def generar_token():
    return secrets.token_hex(32)

# ── Auth decorator ─────────────────────────────────────────────

def require_auth(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        auth = request.headers.get('Authorization', '')
        if not auth.startswith('Bearer '):
            return jsonify({'error': 'No autorizado'}), 401
        token = auth[7:]
        with get_db() as db:
            cur = db.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
            cur.execute('SELECT usuario_id FROM sesiones WHERE token = %s', (token,))
            session = cur.fetchone()
            if not session:
                return jsonify({'error': 'Sesion invalida'}), 401
            request.usuario_id = session['usuario_id']
            cur.execute('SELECT * FROM usuarios WHERE id = %s', (request.usuario_id,))
            request.usuario = cur.fetchone()
        return f(*args, **kwargs)
    return wrapper

# ── Serve static index ────────────────────────────────────────

@app.route('/')
def index():
    return app.send_static_file('index.html')

# ── API: Proyectos (Modo Libre) ────────────────────────────────

@app.route('/api/proyectos', methods=['POST'])
def crear_proyecto():
    data = request.get_json(silent=True) or {}
    usuario = data.get('usuario', '').strip()
    nombre = data.get('nombre', '').strip()

    if not usuario or not nombre:
        return jsonify({'error': 'Nombre de proyecto y usuario son obligatorios'}), 400

    proj_id = generar_id()
    try:
        with get_db() as db:
            cur = db.cursor()
            cur.execute('INSERT INTO proyectos (id, nombre) VALUES (%s, %s)', (proj_id, nombre))
    except Exception as e:
        return jsonify({'error': 'Error al crear proyecto: ' + str(e)}), 500

    return jsonify({'id': proj_id, 'nombre': nombre, 'usuario': usuario}), 201

@app.route('/api/proyectos/join', methods=['POST'])
def unirse_proyecto():
    data = request.get_json(silent=True) or {}
    proj_id = data.get('id', '').strip()
    usuario = data.get('usuario', '').strip()
    nombre = data.get('nombre', '').strip()

    if not proj_id or not usuario or not nombre:
        return jsonify({'error': 'ID, nombre de proyecto y usuario son obligatorios'}), 400

    with get_db() as db:
        cur = db.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute('SELECT * FROM proyectos WHERE id = %s', (proj_id,))
        proyecto = cur.fetchone()
        if not proyecto:
            return jsonify({'error': 'Proyecto no encontrado. Verifica el ID.'}), 404
        if proyecto['nombre'].strip().lower() != nombre.lower():
            return jsonify({'error': 'El nombre del proyecto no coincide con el ID.'}), 400

    return jsonify({'id': proj_id, 'nombre': proyecto['nombre'], 'usuario': usuario})

@app.route('/api/proyectos/<proj_id>', methods=['GET'])
def obtener_proyecto(proj_id):
    with get_db() as db:
        cur = db.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute('SELECT * FROM proyectos WHERE id = %s', (proj_id,))
        proyecto = cur.fetchone()
        if not proyecto:
            return jsonify({'error': 'Proyecto no encontrado'}), 404

        cur.execute(
            'SELECT * FROM rendiciones WHERE proyecto_id = %s ORDER BY creado_en DESC',
            (proj_id,)
        )
        rendiciones = cur.fetchall()

        return jsonify({
            'id': proyecto['id'],
            'nombre': proyecto['nombre'],
            'creado_en': proyecto['creado_en'].isoformat() if proyecto.get('creado_en') else None,
            'rendiciones': [dict(r) for r in rendiciones]
        })

# ── API: Rendiciones (Modo Libre) ──────────────────────────────

@app.route('/api/proyectos/<proj_id>/rendiciones', methods=['POST'])
def agregar_rendicion(proj_id):
    data = request.get_json(silent=True) or {}

    with get_db() as db:
        cur = db.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute('SELECT id FROM proyectos WHERE id = %s', (proj_id,))
        if not cur.fetchone():
            return jsonify({'error': 'Proyecto no encontrado'}), 404

        cur.execute('''INSERT INTO rendiciones (
            proyecto_id, nombre_persona_gasto, boleta_o_factura,
            empresa_emite, nro_boleta_factura, monto_neto, monto_total, fecha, hora, imagen_url
        ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)''', (
            proj_id,
            data.get('nombrePersonaGasto', ''),
            data.get('tipoDocumento', ''),
            data.get('empresaEmite', ''),
            data.get('nroDocumento', '') or '',
            data.get('montoNeto', 0) or 0,
            data.get('montoTotal', 0) or 0,
            data.get('fecha', ''),
            data.get('hora', ''),
            data.get('imagenUrl', '') or ''
        ))

    return jsonify({'ok': True}), 201

# ── API: Excel (Modo Libre) ────────────────────────────────────

@app.route('/api/proyectos/<proj_id>/excel', methods=['GET'])
def descargar_excel(proj_id):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        return jsonify({'error': 'openpyxl no esta instalado en el servidor'}), 500

    with get_db() as db:
        cur = db.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute('SELECT * FROM proyectos WHERE id = %s', (proj_id,))
        proyecto = cur.fetchone()
        if not proyecto:
            return jsonify({'error': 'Proyecto no encontrado'}), 404

        cur.execute(
            'SELECT * FROM rendiciones WHERE proyecto_id = %s ORDER BY creado_en ASC',
            (proj_id,)
        )
        rendiciones = cur.fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = 'Rendiciones'

    headers = ['nombre_plan', 'id_proyecto', 'nombre_persona_gasto', 'boleta_o_factura',
               'empresa_emite', 'nro_boleta_factura', 'monto_neto', 'monto_total']
    header_fill = PatternFill(start_color='4361EE', end_color='4361EE', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font

    for i, r in enumerate(rendiciones, 2):
        ws.cell(row=i, column=1, value=proyecto['nombre'])
        ws.cell(row=i, column=2, value=proj_id)
        ws.cell(row=i, column=3, value=r['nombre_persona_gasto'])
        ws.cell(row=i, column=4, value=r['boleta_o_factura'])
        ws.cell(row=i, column=5, value=r['empresa_emite'])
        ws.cell(row=i, column=6, value=r['nro_boleta_factura'])
        ws.cell(row=i, column=7, value=r['monto_neto'])
        ws.cell(row=i, column=8, value=r['monto_total'])

    ws.column_dimensions['A'].width = 24
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 22
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 28
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 14
    ws.column_dimensions['H'].width = 14

    tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    wb.save(tmp.name)

    nombre_archivo = f"{proyecto['nombre'].replace(' ', '_')}_{proj_id}.xlsx"
    return send_file(tmp.name, as_attachment=True, download_name=nombre_archivo,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

# ═══════════════════════════════════════════════════════════════
#  API: Modo Ejecutivo — Auth
# ═══════════════════════════════════════════════════════════════

@app.route('/api/auth/register', methods=['POST'])
def auth_register():
    data = request.get_json(silent=True) or {}
    email = data.get('email', '').strip().lower()
    password = data.get('password', '').strip()
    rol = data.get('rol', '').strip()

    if not email or not password:
        return jsonify({'error': 'Email y contrasena son obligatorios'}), 400
    if len(password) < 6:
        return jsonify({'error': 'La contrasena debe tener al menos 6 caracteres'}), 400
    if rol not in ('trabajador', 'contador'):
        return jsonify({'error': 'El rol debe ser trabajador o contador'}), 400

    with get_db() as db:
        cur = db.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute('SELECT id FROM usuarios WHERE email = %s', (email,))
        if cur.fetchone():
            return jsonify({'error': 'Este email ya esta registrado'}), 409

        password_hash = generate_password_hash(password)
        codigo = None
        if rol == 'trabajador':
            while True:
                codigo = generar_id()
                cur.execute('SELECT id FROM usuarios WHERE codigo_trabajador = %s', (codigo,))
                if not cur.fetchone():
                    break

        cur.execute(
            'INSERT INTO usuarios (email, password_hash, rol, codigo_trabajador) VALUES (%s, %s, %s, %s) RETURNING id',
            (email, password_hash, rol, codigo)
        )
        usuario_id = cur.fetchone()['id']

    return jsonify({
        'id': usuario_id,
        'email': email,
        'rol': rol,
        'codigo_trabajador': codigo
    }), 201

@app.route('/api/auth/login', methods=['POST'])
def auth_login():
    data = request.get_json(silent=True) or {}
    email = data.get('email', '').strip().lower()
    password = data.get('password', '').strip()

    if not email or not password:
        return jsonify({'error': 'Email y contrasena son obligatorios'}), 400

    with get_db() as db:
        cur = db.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute('SELECT * FROM usuarios WHERE email = %s', (email,))
        usuario = cur.fetchone()
        if not usuario or not check_password_hash(usuario['password_hash'], password):
            return jsonify({'error': 'Email o contrasena incorrectos'}), 401

        token = generar_token()
        cur.execute('INSERT INTO sesiones (token, usuario_id) VALUES (%s, %s)', (token, usuario['id']))

        return jsonify({
            'token': token,
            'usuario': {
                'id': usuario['id'],
                'email': usuario['email'],
                'rol': usuario['rol'],
                'codigo_trabajador': usuario.get('codigo_trabajador')
            }
        })

@app.route('/api/auth/logout', methods=['POST'])
@require_auth
def auth_logout():
    auth = request.headers.get('Authorization', '')
    token = auth[7:]
    with get_db() as db:
        cur = db.cursor()
        cur.execute('DELETE FROM sesiones WHERE token = %s', (token,))
    return jsonify({'ok': True})

@app.route('/api/auth/me', methods=['GET'])
@require_auth
def auth_me():
    return jsonify({
        'id': request.usuario['id'],
        'email': request.usuario['email'],
        'rol': request.usuario['rol'],
        'codigo_trabajador': request.usuario.get('codigo_trabajador')
    })

# ═══════════════════════════════════════════════════════════════
#  API: Modo Ejecutivo — Rendiciones Ejecutivas
# ═══════════════════════════════════════════════════════════════

@app.route('/api/rendiciones', methods=['POST'])
@require_auth
def crear_rendicion_ejecutiva():
    if request.usuario['rol'] != 'trabajador':
        return jsonify({'error': 'Solo los trabajadores pueden crear rendiciones'}), 403

    data = request.get_json(silent=True) or {}
    tipo = data.get('tipo', '').strip()
    nombre = data.get('nombre', '').strip()
    fecha = data.get('fecha', '').strip()
    monto_total = data.get('monto_total', 0) or 0

    if tipo not in ('compania', 'restitucion'):
        return jsonify({'error': 'Tipo de rendicion invalido'}), 400
    if not nombre:
        return jsonify({'error': 'El nombre de la rendicion es obligatorio'}), 400
    if not fecha:
        return jsonify({'error': 'La fecha es obligatoria'}), 400
    if monto_total <= 0:
        return jsonify({'error': 'El monto a rendir debe ser mayor a 0'}), 400

    with get_db() as db:
        cur = db.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute(
            'INSERT INTO rendiciones_ejecutivas (usuario_id, tipo, nombre, fecha, monto_total) VALUES (%s, %s, %s, %s, %s) RETURNING *',
            (request.usuario_id, tipo, nombre, fecha, monto_total)
        )
        rendicion = dict(cur.fetchone())
        rendicion['creado_en'] = rendicion['creado_en'].isoformat() if rendicion.get('creado_en') else None

    return jsonify(rendicion), 201

@app.route('/api/rendiciones', methods=['GET'])
@require_auth
def listar_rendiciones():
    usuario_id = request.usuario_id
    rol = request.usuario['rol']
    estado = request.args.get('estado', '').strip()

    with get_db() as db:
        cur = db.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

        if rol == 'trabajador':
            query = 'SELECT * FROM rendiciones_ejecutivas WHERE usuario_id = %s'
            params = [usuario_id]
            if estado:
                query += ' AND estado = %s'
                params.append(estado)
            query += ' ORDER BY creado_en DESC'
            cur.execute(query, params)
        else:
            # Contador: ver rendiciones de trabajadores vinculados
            query = '''SELECT re.*, u.email as trabajador_email
                       FROM rendiciones_ejecutivas re
                       JOIN vinculaciones v ON re.usuario_id = v.trabajador_id
                       JOIN usuarios u ON re.usuario_id = u.id
                       WHERE v.contador_id = %s'''
            params = [usuario_id]
            if estado:
                query += ' AND re.estado = %s'
                params.append(estado)
            query += ' ORDER BY re.creado_en DESC'
            cur.execute(query, params)

        rendiciones = [dict(r) for r in cur.fetchall()]
        for r in rendiciones:
            if r.get('creado_en'):
                r['creado_en'] = r['creado_en'].isoformat()
            cur.execute(
                'SELECT COALESCE(SUM(monto_total), 0) as total_rendido FROM detalles_rendicion WHERE rendicion_id = %s',
                (r['id'],)
            )
            suma = cur.fetchone()
            r['total_rendido'] = float(suma['total_rendido']) if suma else 0

    return jsonify(rendiciones)

@app.route('/api/rendiciones/<int:rendicion_id>', methods=['GET'])
@require_auth
def obtener_rendicion(rendicion_id):
    with get_db() as db:
        cur = db.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute('SELECT * FROM rendiciones_ejecutivas WHERE id = %s', (rendicion_id,))
        rendicion = cur.fetchone()
        if not rendicion:
            return jsonify({'error': 'Rendicion no encontrada'}), 404

        if request.usuario['rol'] == 'trabajador' and rendicion['usuario_id'] != request.usuario_id:
            return jsonify({'error': 'No autorizado'}), 403

        if request.usuario['rol'] == 'contador':
            cur.execute(
                'SELECT id FROM vinculaciones WHERE contador_id = %s AND trabajador_id = %s',
                (request.usuario_id, rendicion['usuario_id'])
            )
            if not cur.fetchone():
                return jsonify({'error': 'No autorizado'}), 403

        rendicion = dict(rendicion)
        if rendicion.get('creado_en'):
            rendicion['creado_en'] = rendicion['creado_en'].isoformat()

        cur.execute(
            'SELECT * FROM detalles_rendicion WHERE rendicion_id = %s ORDER BY creado_en DESC',
            (rendicion_id,)
        )
        detalles = [dict(d) for d in cur.fetchall()]
        for d in detalles:
            if d.get('creado_en'):
                d['creado_en'] = d['creado_en'].isoformat()

        rendicion['detalles'] = detalles

        cur.execute(
            'SELECT COALESCE(SUM(monto_total), 0) as total_rendido FROM detalles_rendicion WHERE rendicion_id = %s',
            (rendicion_id,)
        )
        suma = cur.fetchone()
        rendicion['total_rendido'] = float(suma['total_rendido']) if suma else 0

    return jsonify(rendicion)

@app.route('/api/rendiciones/<int:rendicion_id>/cerrar', methods=['PUT'])
@require_auth
def cerrar_rendicion(rendicion_id):
    if request.usuario['rol'] != 'trabajador':
        return jsonify({'error': 'Solo los trabajadores pueden cerrar rendiciones'}), 403

    with get_db() as db:
        cur = db.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute('SELECT * FROM rendiciones_ejecutivas WHERE id = %s', (rendicion_id,))
        rendicion = cur.fetchone()
        if not rendicion:
            return jsonify({'error': 'Rendicion no encontrada'}), 404
        if rendicion['usuario_id'] != request.usuario_id:
            return jsonify({'error': 'No autorizado'}), 403
        if rendicion['estado'] == 'cerrada':
            return jsonify({'error': 'La rendicion ya esta cerrada'}), 400

        cur.execute('UPDATE rendiciones_ejecutivas SET estado = %s WHERE id = %s', ('cerrada', rendicion_id))

    return jsonify({'ok': True, 'estado': 'cerrada'})

# ═══════════════════════════════════════════════════════════════
#  API: Modo Ejecutivo — Detalles de Rendicion
# ═══════════════════════════════════════════════════════════════

@app.route('/api/rendiciones/<int:rendicion_id>/detalles', methods=['POST'])
@require_auth
def agregar_detalle(rendicion_id):
    if request.usuario['rol'] != 'trabajador':
        return jsonify({'error': 'Solo los trabajadores pueden agregar detalles'}), 403

    data = request.get_json(silent=True) or {}

    with get_db() as db:
        cur = db.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute('SELECT * FROM rendiciones_ejecutivas WHERE id = %s', (rendicion_id,))
        rendicion = cur.fetchone()
        if not rendicion:
            return jsonify({'error': 'Rendicion no encontrada'}), 404
        if rendicion['usuario_id'] != request.usuario_id:
            return jsonify({'error': 'No autorizado'}), 403
        if rendicion['estado'] == 'cerrada':
            return jsonify({'error': 'No se pueden agregar detalles a una rendicion cerrada'}), 400

        tipo_gasto_entrada = data.get('tipo_gasto_entrada', '').strip()
        nro_documento = data.get('nro_documento', '').strip()

        if tipo_gasto_entrada not in ('boleta', 'factura', 'devolucion', 'manual'):
            return jsonify({'error': 'Tipo de gasto invalido'}), 400

        if nro_documento:
            cur.execute(
                'SELECT id FROM detalles_rendicion WHERE rendicion_id = %s AND nro_documento = %s',
                (rendicion_id, nro_documento)
            )
            if cur.fetchone():
                return jsonify({'error': 'El numero de documento ya existe en esta rendicion'}), 409

        fecha = data.get('fecha', '').strip()
        rut_emisor = data.get('rut_emisor', '').strip()
        monto_total = data.get('monto_total', 0) or 0
        monto_neto = data.get('monto_neto', 0) or 0
        monto_iva = data.get('monto_iva', 0) or 0

        if monto_total > 0 and monto_neto <= 0:
            monto_neto = round(monto_total / 1.19)
            monto_iva = monto_total - monto_neto
        empresa_emite = data.get('empresa_emite', '').strip()
        tipo_gasto = data.get('tipo_gasto', '').strip()
        descripcion = data.get('descripcion', '').strip()
        imagen_url = data.get('imagen_url', '').strip()
        centro_costo_id = data.get('centro_costo_id')

        cur.execute('''INSERT INTO detalles_rendicion (
            rendicion_id, tipo_gasto_entrada, fecha, rut_emisor, nro_documento,
            monto_total, monto_neto, monto_iva, empresa_emite, tipo_gasto, descripcion, imagen_url, centro_costo_id
        ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s) RETURNING *''', (
            rendicion_id, tipo_gasto_entrada, fecha, rut_emisor, nro_documento,
            monto_total, monto_neto, monto_iva, empresa_emite, tipo_gasto, descripcion, imagen_url,
            int(centro_costo_id) if centro_costo_id and centro_costo_id != '' and centro_costo_id != 'null' else None
        ))
        detalle = dict(cur.fetchone())
        if detalle.get('creado_en'):
            detalle['creado_en'] = detalle['creado_en'].isoformat()

    return jsonify(detalle), 201

@app.route('/api/rendiciones/<int:rendicion_id>/detalles/<int:detalle_id>', methods=['PUT'])
@require_auth
def actualizar_detalle(rendicion_id, detalle_id):
    if request.usuario['rol'] != 'trabajador':
        return jsonify({'error': 'Solo los trabajadores pueden modificar detalles'}), 403

    data = request.get_json(silent=True) or {}

    with get_db() as db:
        cur = db.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute('SELECT * FROM rendiciones_ejecutivas WHERE id = %s', (rendicion_id,))
        rendicion = cur.fetchone()
        if not rendicion or rendicion['usuario_id'] != request.usuario_id:
            return jsonify({'error': 'No autorizado'}), 403
        if rendicion['estado'] == 'cerrada':
            return jsonify({'error': 'No se pueden modificar detalles de una rendicion cerrada'}), 400

        cur.execute('SELECT * FROM detalles_rendicion WHERE id = %s AND rendicion_id = %s', (detalle_id, rendicion_id))
        detalle = cur.fetchone()
        if not detalle:
            return jsonify({'error': 'Detalle no encontrado'}), 404

        nro_documento = data.get('nro_documento', '').strip()
        if nro_documento and nro_documento != detalle['nro_documento']:
            cur.execute(
                'SELECT id FROM detalles_rendicion WHERE rendicion_id = %s AND nro_documento = %s AND id != %s',
                (rendicion_id, nro_documento, detalle_id)
            )
            if cur.fetchone():
                return jsonify({'error': 'El numero de documento ya existe en esta rendicion'}), 409

        campos = ['fecha', 'rut_emisor', 'nro_documento', 'monto_total', 'monto_neto',
                   'monto_iva', 'empresa_emite', 'tipo_gasto', 'descripcion', 'imagen_url']
        sets = []
        params = []
        for campo in campos:
            if campo in data:
                sets.append(campo + ' = %s')
                params.append(data[campo])
        if not sets:
            return jsonify({'error': 'No hay campos para actualizar'}), 400
        params.extend([detalle_id, rendicion_id])
        cur.execute(
            'UPDATE detalles_rendicion SET ' + ', '.join(sets) + ' WHERE id = %s AND rendicion_id = %s',
            params
        )

    return jsonify({'ok': True})

@app.route('/api/rendiciones/<int:rendicion_id>/detalles/<int:detalle_id>', methods=['DELETE'])
@require_auth
def eliminar_detalle(rendicion_id, detalle_id):
    if request.usuario['rol'] != 'trabajador':
        return jsonify({'error': 'Solo los trabajadores pueden eliminar detalles'}), 403

    with get_db() as db:
        cur = db.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute('SELECT * FROM rendiciones_ejecutivas WHERE id = %s', (rendicion_id,))
        rendicion = cur.fetchone()
        if not rendicion or rendicion['usuario_id'] != request.usuario_id:
            return jsonify({'error': 'No autorizado'}), 403
        if rendicion['estado'] == 'cerrada':
            return jsonify({'error': 'No se pueden eliminar detalles de una rendicion cerrada'}), 400

        cur.execute('DELETE FROM detalles_rendicion WHERE id = %s AND rendicion_id = %s', (detalle_id, rendicion_id))

    return jsonify({'ok': True})

# ═══════════════════════════════════════════════════════════════
#  API: Modo Ejecutivo — Vinculaciones (Contador)
# ═══════════════════════════════════════════════════════════════

@app.route('/api/vincular', methods=['POST'])
@require_auth
def vincular_trabajador():
    if request.usuario['rol'] != 'contador':
        return jsonify({'error': 'Solo los contadores pueden vincular trabajadores'}), 403

    data = request.get_json(silent=True) or {}
    codigo = data.get('codigo', '').strip()

    if not codigo:
        return jsonify({'error': 'El codigo de trabajador es obligatorio'}), 400

    with get_db() as db:
        cur = db.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute('SELECT * FROM usuarios WHERE codigo_trabajador = %s AND rol = %s', (codigo, 'trabajador'))
        trabajador = cur.fetchone()
        if not trabajador:
            return jsonify({'error': 'Trabajador no encontrado. Verifica el codigo.'}), 404

        cur.execute(
            'SELECT id FROM vinculaciones WHERE contador_id = %s AND trabajador_id = %s',
            (request.usuario_id, trabajador['id'])
        )
        if cur.fetchone():
            return jsonify({'error': 'Ya estas vinculado a este trabajador'}), 409

        cur.execute(
            'INSERT INTO vinculaciones (contador_id, trabajador_id) VALUES (%s, %s) RETURNING id',
            (request.usuario_id, trabajador['id'])
        )

    return jsonify({
        'ok': True,
        'trabajador': {'id': trabajador['id'], 'email': trabajador['email']}
    }), 201

@app.route('/api/contador/trabajadores', methods=['GET'])
@require_auth
def listar_trabajadores():
    if request.usuario['rol'] != 'contador':
        return jsonify({'error': 'Solo los contadores pueden ver sus trabajadores'}), 403

    with get_db() as db:
        cur = db.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute('''SELECT u.id, u.email, u.codigo_trabajador, v.creado_en as vinculado_en
                       FROM usuarios u
                       JOIN vinculaciones v ON u.id = v.trabajador_id
                       WHERE v.contador_id = %s
                       ORDER BY v.creado_en DESC''', (request.usuario_id,))
        trabajadores = [dict(t) for t in cur.fetchall()]
        for t in trabajadores:
            if t.get('vinculado_en'):
                t['vinculado_en'] = t['vinculado_en'].isoformat()

    return jsonify(trabajadores)

# ═══════════════════════════════════════════════════════════════
#  API: Centros de Costo
# ═══════════════════════════════════════════════════════════════

@app.route('/api/centros-costo', methods=['GET'])
@require_auth
def listar_centros_costo():
    with get_db() as db:
        cur = db.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute('SELECT * FROM centros_costo ORDER BY nombre ASC')
        centros = [dict(c) for c in cur.fetchall()]
        for c in centros:
            if c.get('creado_en'):
                c['creado_en'] = c['creado_en'].isoformat()
    return jsonify(centros)

@app.route('/api/centros-costo', methods=['POST'])
@require_auth
def crear_centro_costo():
    if request.usuario['rol'] != 'contador':
        return jsonify({'error': 'Solo los contadores pueden crear centros de costo'}), 403

    data = request.get_json(silent=True) or {}
    codigo = data.get('codigo', '').strip()
    nombre = data.get('nombre', '').strip()
    if not codigo or not nombre:
        return jsonify({'error': 'Codigo y nombre son obligatorios'}), 400

    with get_db() as db:
        cur = db.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute('SELECT id FROM centros_costo WHERE codigo = %s', (codigo,))
        if cur.fetchone():
            return jsonify({'error': 'Este codigo ya existe'}), 409
        cur.execute('INSERT INTO centros_costo (codigo, nombre) VALUES (%s, %s) RETURNING *', (codigo, nombre))
        centro = dict(cur.fetchone())
        if centro.get('creado_en'):
            centro['creado_en'] = centro['creado_en'].isoformat()
    return jsonify(centro), 201

@app.route('/api/centros-costo/<int:centro_id>', methods=['DELETE'])
@require_auth
def eliminar_centro_costo(centro_id):
    if request.usuario['rol'] != 'contador':
        return jsonify({'error': 'Solo los contadores pueden eliminar centros de costo'}), 403

    with get_db() as db:
        cur = db.cursor()
        cur.execute('UPDATE detalles_rendicion SET centro_costo_id = NULL WHERE centro_costo_id = %s', (centro_id,))
        cur.execute('DELETE FROM centros_costo WHERE id = %s', (centro_id,))
    return jsonify({'ok': True})

@app.route('/api/rendiciones/<int:rendicion_id>/detalles/<int:detalle_id>/centro-costo', methods=['PUT'])
@require_auth
def asignar_centro_costo(rendicion_id, detalle_id):
    if request.usuario['rol'] not in ('trabajador', 'contador'):
        return jsonify({'error': 'Rol no autorizado'}), 403

    data = request.get_json(silent=True) or {}
    centro_id = data.get('centro_costo_id')

    with get_db() as db:
        cur = db.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute('SELECT * FROM rendiciones_ejecutivas WHERE id = %s', (rendicion_id,))
        rendicion = cur.fetchone()
        if not rendicion:
            return jsonify({'error': 'Rendicion no encontrada'}), 404

        is_owner = (request.usuario_id == rendicion['usuario_id'])
        is_contador = request.usuario['rol'] == 'contador'
        if not is_owner and not is_contador:
            return jsonify({'error': 'No autorizado'}), 403
        if is_contador and not is_owner:
            cur.execute(
                'SELECT id FROM vinculaciones WHERE contador_id = %s AND trabajador_id = %s',
                (request.usuario_id, rendicion['usuario_id'])
            )
            if not cur.fetchone():
                return jsonify({'error': 'No estas vinculado a este trabajador'}), 403

        cur.execute('SELECT id FROM detalles_rendicion WHERE id = %s AND rendicion_id = %s', (detalle_id, rendicion_id))
        if not cur.fetchone():
            return jsonify({'error': 'Detalle no encontrado'}), 404

        if centro_id is not None:
            if centro_id != '' and centro_id != 'null':
                cur.execute('SELECT id FROM centros_costo WHERE id = %s', (centro_id,))
                if not cur.fetchone():
                    return jsonify({'error': 'Centro de costo no encontrado'}), 404
                cur.execute('UPDATE detalles_rendicion SET centro_costo_id = %s WHERE id = %s', (centro_id, detalle_id))
            else:
                cur.execute('UPDATE detalles_rendicion SET centro_costo_id = NULL WHERE id = %s', (detalle_id,))
        else:
            cur.execute('UPDATE detalles_rendicion SET centro_costo_id = NULL WHERE id = %s', (detalle_id,))

    return jsonify({'ok': True})

# ═══════════════════════════════════════════════════════════════
#  API: Excel Contable
# ═══════════════════════════════════════════════════════════════

@app.route('/api/rendiciones/<int:rendicion_id>/excel', methods=['GET'])
@require_auth
def descargar_excel_contable(rendicion_id):
    with get_db() as db:
        cur = db.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute('SELECT * FROM rendiciones_ejecutivas WHERE id = %s', (rendicion_id,))
        rendicion = cur.fetchone()
        if not rendicion:
            return jsonify({'error': 'Rendicion no encontrada'}), 404

        can_access = (request.usuario_id == rendicion['usuario_id'])
        if not can_access and request.usuario['rol'] == 'contador':
            cur.execute(
                'SELECT id FROM vinculaciones WHERE contador_id = %s AND trabajador_id = %s',
                (request.usuario_id, rendicion['usuario_id'])
            )
            can_access = cur.fetchone() is not None
        if not can_access:
            return jsonify({'error': 'No autorizado'}), 403

        cur.execute(
            'SELECT d.*, cc.nombre as centro_costo_nombre, cc.codigo as centro_costo_codigo FROM detalles_rendicion d LEFT JOIN centros_costo cc ON d.centro_costo_id = cc.id WHERE d.rendicion_id = %s ORDER BY d.creado_en ASC',
            (rendicion_id,)
        )
        detalles = [dict(d) for d in cur.fetchall()]

        cur.execute('SELECT COALESCE(SUM(monto_total), 0) as total FROM detalles_rendicion WHERE rendicion_id = %s', (rendicion_id,))
        total_render = float(cur.fetchone()['total'])

    try:
        es_compania = rendicion['tipo'] == 'compania'
        monto_total = float(rendicion['monto_total'])
        saldo_sobrante = monto_total - total_render
        if saldo_sobrante < 0:
            saldo_sobrante = 0

        wb, ws = _build_excel_rows(rendicion, detalles, es_compania, monto_total, saldo_sobrante, total_render)
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
    except Exception as e:
        return jsonify({'error': 'Error al generar Excel: ' + str(e)}), 500

    nombre_archivo = rendicion['nombre'].replace(' ', '_') + '.xlsx'
    return send_file(output, as_attachment=True, download_name=nombre_archivo,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

# ═══════════════════════════════════════════════════════════════
#  API: Enviar por correo
# ═══════════════════════════════════════════════════════════════

@app.route('/api/rendiciones/<int:rendicion_id>/enviar', methods=['POST'])
@require_auth
def enviar_rendicion_correo(rendicion_id):
    SENDGRID_API_KEY = os.environ.get('SENDGRID_API_KEY', '')
    EMAIL_REMITENTE = os.environ.get('EMAIL_REMITENTE', '')

    if not SENDGRID_API_KEY:
        return jsonify({'error': 'SENDGRID_API_KEY no configurada en el servidor'}), 500
    if not EMAIL_REMITENTE:
        return jsonify({'error': 'EMAIL_REMITENTE no configurado en el servidor'}), 500

    data = request.get_json(silent=True) or {}
    destinatario = data.get('email', '').strip()
    if not destinatario:
        return jsonify({'error': 'El email del destinatario es obligatorio'}), 400

    with get_db() as db:
        cur = db.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute('SELECT * FROM rendiciones_ejecutivas WHERE id = %s', (rendicion_id,))
        rendicion = cur.fetchone()
        if not rendicion:
            return jsonify({'error': 'Rendicion no encontrada'}), 404

        can_access = (request.usuario_id == rendicion['usuario_id'])
        if not can_access and request.usuario['rol'] == 'contador':
            cur.execute(
                'SELECT id FROM vinculaciones WHERE contador_id = %s AND trabajador_id = %s',
                (request.usuario_id, rendicion['usuario_id'])
            )
            can_access = cur.fetchone() is not None
        if not can_access:
            return jsonify({'error': 'No autorizado'}), 403

        cur.execute(
            'SELECT d.*, cc.nombre as centro_costo_nombre, cc.codigo as centro_costo_codigo FROM detalles_rendicion d LEFT JOIN centros_costo cc ON d.centro_costo_id = cc.id WHERE d.rendicion_id = %s ORDER BY d.creado_en ASC',
            (rendicion_id,)
        )
        detalles = [dict(d) for d in cur.fetchall()]

        cur.execute('SELECT COALESCE(SUM(monto_total), 0) as total FROM detalles_rendicion WHERE rendicion_id = %s', (rendicion_id,))
        total_render = float(cur.fetchone()['total'])

    try:
        es_compania = rendicion['tipo'] == 'compania'
        monto_total = float(rendicion['monto_total'])
        saldo_sobrante = monto_total - total_render
        if saldo_sobrante < 0:
            saldo_sobrante = 0

        wb, ws = _build_excel_rows(rendicion, detalles, es_compania, monto_total, saldo_sobrante, total_render)
        excel_io = io.BytesIO()
        wb.save(excel_io)
        excel_io.seek(0)
        excel_b64 = base64.b64encode(excel_io.read()).decode('utf-8')

        tipo_label = 'Dinero entregado por la compania' if es_compania else 'Restitucion fondos propios'
        fotos_html_parts = []
        for d in detalles:
            if d.get('imagen_url'):
                label = str(d.get('empresa_emite') or d['tipo_gasto_entrada'])
                desc = d.get('descripcion', '')
                if desc:
                    label += ' - ' + str(desc)
                url = str(d['imagen_url'])
                fotos_html_parts.append('<p>' + label + '<br><a href="' + url + '">Ver foto</a></p>')
        fotos_html = '\n'.join(fotos_html_parts)

        html = '<html><body style="font-family:Arial,sans-serif">'
        html += '<h2 style="color:#4361EE">Rendicion: ' + str(rendicion['nombre']) + '</h2>'
        html += '<p><strong>Tipo:</strong> ' + tipo_label + '</p>'
        html += '<p><strong>Fecha:</strong> ' + str(rendicion['fecha']) + '</p>'
        html += '<p><strong>Monto a rendir:</strong> $' + format(int(monto_total), ',d') + '</p>'
        html += '<p><strong>Total rendido:</strong> $' + format(int(total_render), ',d') + '</p>'
        html += '<p><strong>Saldo:</strong> $' + format(int(saldo_sobrante), ',d') + '</p>'
        html += fotos_html
        html += '<p style="color:#666;font-size:12px;margin-top:24px">Enviado desde Cuentas Claras</p>'
        html += '</body></html>'

        nombre_archivo = str(rendicion['nombre']).replace(' ', '_') + '.xlsx'

        payload = json.dumps({
            'personalizations': [{'to': [{'email': destinatario}]}],
            'from': {'email': EMAIL_REMITENTE, 'name': 'Cuentas Claras'},
            'subject': 'Rendicion: ' + str(rendicion['nombre']) + ' - ' + str(rendicion['fecha']),
            'content': [{'type': 'text/html', 'value': html}],
            'attachments': [{
                'filename': nombre_archivo,
                'content': excel_b64,
                'type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            }]
        }).encode('utf-8')

        req = urllib.request.Request(
            'https://api.sendgrid.com/v3/mail/send',
            data=payload,
            headers={
                'Authorization': 'Bearer ' + SENDGRID_API_KEY,
                'Content-Type': 'application/json'
            },
            method='POST'
        )

        with urllib.request.urlopen(req, timeout=30) as resp:
            resp_data = resp.read().decode('utf-8')

    except Exception as e:
        msg = str(e)
        return jsonify({'error': 'Error al enviar el correo: ' + msg}), 500

    return jsonify({'ok': True})

def _build_excel_rows(rendicion, detalles, es_compania, monto_total, saldo_sobrante, total_render):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = 'Rendicion'

    h_fill = PatternFill(start_color='4361EE', end_color='4361EE', fill_type='solid')
    h_font = Font(color='FFFFFF', bold=True, size=11)
    thin_border = Border(
        left=Side(style='thin', color='CCCCCC'), right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'), bottom=Side(style='thin', color='CCCCCC')
    )
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell_align = Alignment(vertical='center', wrap_text=True)
    number_align = Alignment(horizontal='right', vertical='center')

    headers = ['Fecha', 'Tipo Documento', 'Nro Documento', 'Cuenta Contable',
               'Descripcion', 'Centro de Costo', 'Debe', 'Haber']
    col_widths = [14, 20, 16, 22, 40, 20, 16, 16]

    for col_idx, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = h_fill
        cell.font = h_font
        cell.alignment = header_align
        cell.border = thin_border
        ws.column_dimensions[cell.column_letter].width = w

    row = [2]

    def wr(fecha, tipo_doc, nro_doc, cuenta, descripcion, centro_costo, debe, haber):
        values = [fecha, tipo_doc, nro_doc, cuenta, descripcion, centro_costo or '', debe, haber]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row[0], column=col_idx, value=val)
            cell.border = thin_border
            cell.alignment = number_align if col_idx in (7, 8) and val else cell_align
            if col_idx in (7, 8) and isinstance(val, (int, float)):
                cell.number_format = '#,##0'
                cell.font = Font(color='1e7e34' if col_idx == 7 else 'b02a37', size=10)
        row[0] += 1

    if es_compania:
        desc_fondo = 'Fondo entregado: ' + rendicion['nombre']
        wr(rendicion['fecha'], 'Fondo por rendir', '', 'Activo por rendir', desc_fondo, '', monto_total, 0)
        wr(rendicion['fecha'], 'Fondo por rendir', '', 'Banco', desc_fondo, '', 0, monto_total)

    for d in detalles:
        tipo_doc = d['tipo_gasto_entrada'].capitalize()
        if d['tipo_gasto_entrada'] == 'devolucion':
            tipo_doc = 'Dev. dinero'
        nro = d['nro_documento'] or ''
        desc = d['descripcion'] or d['empresa_emite'] or ''
        fecha = d['fecha'] or rendicion['fecha']
        monto = float(d['monto_total'] or 0)
        cc_nombre = d.get('centro_costo_nombre', '')
        cc_codigo = d.get('centro_costo_codigo', '')
        cc = (cc_codigo + ' - ' + cc_nombre) if cc_codigo and cc_nombre else (cc_codigo or cc_nombre)

        if es_compania:
            if d['tipo_gasto_entrada'] == 'devolucion':
                wr(fecha, 'Baucher bancario', nro, 'Banco', desc, cc, monto, 0)
                wr(fecha, 'Baucher bancario', nro, 'Activo por rendir', desc, cc, 0, monto)
            else:
                wr(fecha, tipo_doc, nro, 'Gasto', desc, cc, monto, 0)
                wr(fecha, tipo_doc, nro, 'Activo por rendir', desc, cc, 0, monto)
        else:
            if d['tipo_gasto_entrada'] == 'devolucion':
                wr(fecha, 'Baucher bancario', nro, 'Cuentas x pagar', desc, cc, monto, 0)
                wr(fecha, 'Baucher bancario', nro, 'Gasto', desc, cc, 0, monto)
            else:
                wr(fecha, tipo_doc, nro, 'Gasto', desc, cc, monto, 0)
                wr(fecha, tipo_doc, nro, 'Cuentas x pagar', desc, cc, 0, monto)

    if es_compania and saldo_sobrante > 0:
        wr(rendicion['fecha'], 'Baucher bancario', '', 'Banco', 'Devolucion fondo sobrante', '', saldo_sobrante, 0)
        wr(rendicion['fecha'], 'Baucher bancario', '', 'Activo por rendir', 'Devolucion fondo sobrante', '', 0, saldo_sobrante)

    row[0] += 1
    total_cell = ws.cell(row=row[0], column=6, value='TOTAL')
    total_cell.font = Font(bold=True, size=11)
    total_cell.alignment = Alignment(horizontal='right', vertical='center')
    total_cell.border = thin_border

    for c in (ws.cell(row=row[0], column=7), ws.cell(row=row[0], column=8)):
        c.font = Font(bold=True, size=11)
        c.alignment = Alignment(horizontal='right', vertical='center')
        c.border = thin_border
        c.number_format = '#,##0'

    return wb, ws

# ── Health check ──────────────────────────────────────────────

@app.route('/api/health')
def health():
    return jsonify({'status': 'ok'})

# ── Start ─────────────────────────────────────────────────────

init_db()

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=False)
