import os
import sqlite3
import string
import random
import tempfile
from datetime import datetime
from flask import Flask, request, jsonify, send_file

app = Flask(__name__, static_folder='static', static_url_path='')

DATABASE = 'cuentas.db'

def get_db():
    db = sqlite3.connect(DATABASE)
    db.row_factory = sqlite3.Row
    db.execute("PRAGMA journal_mode=WAL")
    db.execute("PRAGMA foreign_keys=ON")
    return db

def init_db():
    with get_db() as db:
        db.execute('''CREATE TABLE IF NOT EXISTS proyectos (
            id TEXT PRIMARY KEY,
            nombre TEXT NOT NULL,
            creado_en TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )''')
        db.execute('''CREATE TABLE IF NOT EXISTS rendiciones (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            proyecto_id TEXT NOT NULL,
            nombre_persona_gasto TEXT NOT NULL,
            boleta_o_factura TEXT NOT NULL,
            empresa_emite TEXT NOT NULL,
            nro_boleta_factura TEXT,
            monto_neto REAL DEFAULT 0,
            monto_total REAL DEFAULT 0,
            fecha TEXT,
            hora TEXT,
            creado_en TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (proyecto_id) REFERENCES proyectos(id)
        )''')

def generar_id(length=8):
    chars = string.ascii_lowercase + string.digits
    return ''.join(random.choice(chars) for _ in range(length))

# ── Serve static index ────────────────────────────────────────

@app.route('/')
def index():
    return app.send_static_file('index.html')

# ── API: Proyectos ────────────────────────────────────────────

@app.route('/api/proyectos', methods=['POST'])
def crear_proyecto():
    data = request.get_json(silent=True) or {}
    usuario = data.get('usuario', '').strip()
    nombre = data.get('nombre', '').strip()

    if not usuario or not nombre:
        return jsonify({'error': 'Nombre de proyecto y usuario son obligatorios'}), 400

    proj_id = generar_id()
    with get_db() as db:
        db.execute('INSERT INTO proyectos (id, nombre) VALUES (?, ?)', (proj_id, nombre))

    return jsonify({'id': proj_id, 'nombre': nombre, 'usuario': usuario}), 201

@app.route('/api/proyectos/join', methods=['POST'])
def unirse_proyecto():
    data = request.get_json(silent=True) or {}
    proj_id = data.get('id', '').strip()
    usuario = data.get('usuario', '').strip()
    nombre = data.get('nombre', '').strip()

    if not proj_id or not usuario or not nombre:
        return jsonify({'error': 'ID, nombre de proyecto y usuario son obligatorios'}), 400

    with get_db() as db:
        proyecto = db.execute('SELECT * FROM proyectos WHERE id = ?', (proj_id,)).fetchone()
        if not proyecto:
            return jsonify({'error': 'Proyecto no encontrado. Verifica el ID.'}), 404
        if proyecto['nombre'].strip().lower() != nombre.lower():
            return jsonify({'error': 'El nombre del proyecto no coincide con el ID.'}), 400

    return jsonify({'id': proj_id, 'nombre': proyecto['nombre'], 'usuario': usuario})

@app.route('/api/proyectos/<proj_id>', methods=['GET'])
def obtener_proyecto(proj_id):
    with get_db() as db:
        proyecto = db.execute('SELECT * FROM proyectos WHERE id = ?', (proj_id,)).fetchone()
        if not proyecto:
            return jsonify({'error': 'Proyecto no encontrado'}), 404

        rendiciones = db.execute(
            'SELECT * FROM rendiciones WHERE proyecto_id = ? ORDER BY creado_en DESC',
            (proj_id,)
        ).fetchall()

        return jsonify({
            'id': proyecto['id'],
            'nombre': proyecto['nombre'],
            'creado_en': proyecto['creado_en'],
            'rendiciones': [dict(r) for r in rendiciones]
        })

# ── API: Rendiciones ──────────────────────────────────────────

@app.route('/api/proyectos/<proj_id>/rendiciones', methods=['POST'])
def agregar_rendicion(proj_id):
    data = request.get_json(silent=True) or {}

    with get_db() as db:
        proyecto = db.execute('SELECT id FROM proyectos WHERE id = ?', (proj_id,)).fetchone()
        if not proyecto:
            return jsonify({'error': 'Proyecto no encontrado'}), 404

        db.execute('''INSERT INTO rendiciones (
            proyecto_id, nombre_persona_gasto, boleta_o_factura,
            empresa_emite, nro_boleta_factura, monto_neto, monto_total, fecha, hora
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)''', (
            proj_id,
            data.get('nombrePersonaGasto', ''),
            data.get('tipoDocumento', ''),
            data.get('empresaEmite', ''),
            data.get('nroDocumento', '') or '',
            data.get('montoNeto', 0) or 0,
            data.get('montoTotal', 0) or 0,
            data.get('fecha', ''),
            data.get('hora', '')
        ))

    return jsonify({'ok': True}), 201

# ── API: Excel ────────────────────────────────────────────────

@app.route('/api/proyectos/<proj_id>/excel', methods=['GET'])
def descargar_excel(proj_id):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        return jsonify({'error': 'openpyxl no esta instalado en el servidor'}), 500

    with get_db() as db:
        proyecto = db.execute('SELECT * FROM proyectos WHERE id = ?', (proj_id,)).fetchone()
        if not proyecto:
            return jsonify({'error': 'Proyecto no encontrado'}), 404

        rendiciones = db.execute(
            'SELECT * FROM rendiciones WHERE proyecto_id = ? ORDER BY creado_en ASC',
            (proj_id,)
        ).fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = 'Rendiciones'

    headers = ['nombre_plan', 'id_proyecto', 'nombre_persona_gasto', 'boleta_o_factura',
               'empresa_emite', 'nro_boleta_factura', 'monto_neto', 'monto_total']
    header_fill = PatternFill(start_color='4361EE', end_color='4361EE', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font

    for i, r in enumerate(rendiciones, 2):
        ws.cell(row=i, column=1, value=proyecto['nombre'])
        ws.cell(row=i, column=2, value=proj_id)
        ws.cell(row=i, column=3, value=r['nombre_persona_gasto'])
        ws.cell(row=i, column=4, value=r['boleta_o_factura'])
        ws.cell(row=i, column=5, value=r['empresa_emite'])
        ws.cell(row=i, column=6, value=r['nro_boleta_factura'])
        ws.cell(row=i, column=7, value=r['monto_neto'])
        ws.cell(row=i, column=8, value=r['monto_total'])

    ws.column_dimensions['A'].width = 24
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 22
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 28
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 14
    ws.column_dimensions['H'].width = 14

    tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    wb.save(tmp.name)

    nombre_archivo = f"{proyecto['nombre'].replace(' ', '_')}_{proj_id}.xlsx"
    return send_file(tmp.name, as_attachment=True, download_name=nombre_archivo,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

# ── Health check ──────────────────────────────────────────────

@app.route('/api/health')
def health():
    return jsonify({'status': 'ok'})

# ── Start ─────────────────────────────────────────────────────

if __name__ == '__main__':
    init_db()
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=False)
